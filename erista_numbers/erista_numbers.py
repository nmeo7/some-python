from openpyxl import load_workbook, Workbook
import json

# ranges printing starting from a row below

def char_range(c1, c2):
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)
        
wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/C I C 31 08 2021.xlsx", data_only=True)
sheet_ranges = wb['Curency in circulation']
previousDay = sheet_ranges[ 'B72' ]

wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/C I C 31 08 2021.xlsx")
sheet_ranges = wb['Curency in circulation']

out = {}

for x in range(1, 77):
    xx = {}
    for y in char_range ('A', 'N'):
        xx[y] = sheet_ranges[ y + str(x)]
    out[str(x)] = xx

out['73']['B'] = previousDay
out['1']['F'].value = 'CURRENCY IN CIRCULATION ON SEPTEMBER 01,  2021'
out['72']['A'].value = 'Net Issue on 01/09/2021'

def retrieve_range(st, end):
    ret = {}
    for x in range (st, end + 1):
        s = str(sheet_ranges[ 'C' + str(x) ].value)
        if s in series:
            ret[s] = []

            for y in  char_range('D', 'M'):
                n = sheet_ranges[ y + str(x)].value
                if n is None:
                    n = 0
                else:
                    n = float (n)

                ret[s].append (n)
    return ret

def retrieve_range_branch(st, end, lstart = 'C', lend = 'L', header = '4'):
    ret = {}
    for x in char_range(lstart, lend):
        s = str(sheet_ranges[ x + header ].value)
        ret[s] = 0
        
        for y in range (st, end + 1):
            n = sheet_ranges[ x + str(y)].value
            if n is None:
                n = 0
            else:
                n = float (n)

            ret[s] += n
    ret2 = []
    for x in ret:
        ret2.append (ret[x])

    return ret2

def retrieve_line_hq(line):
    ret = []
    for x in char_range('C', 'L'):
        n = str(sheet_ranges[ x + str(line) ].value)

        try:
            n = float(n)
        except:
            n = 0

        ret.append(n)
    return ret

def retrieve_column_hq (col):
    ret = [0]
    for x in range (36, 27, -1):
        n = str(sheet_ranges[ col + str(x) ].value)

        try:
            n = float(n)
        except:
            n = 0

        ret.append(n)
    return ret


def retrieve_laila ():
    ret = []
    for x in range(19, 25):
        n = str(sheet_ranges[ 'D' + str(x) ].value)
        try:
            n = float(n)
        except:
            n = 0
        ret.append(n)

    for x in range(6, 10):
        n = str(sheet_ranges[ 'D' + str(x) ].value)
        try:
            n = float(n)
        except:
            n = 0
        ret.append(n)

    return ret

def retrieve_line (l):
    return retrieve_range (l, l)


wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/C I C 31 08 2021.xlsx", data_only = True)
sheet_ranges = wb['Curency in circulation']

series = ( '2003-2004', '2007', '2008', '2009', '2011', '2013', '2014', '2015', '2019' )
denomination = [ 1, 5, 10, 20, 50, 100, 500, 1000, 2000, 5000 ]

printed = retrieve_range (4, 12)
uncirculated = retrieve_range (14, 23)
crushed = retrieve_range (25, 32)
net = retrieve_range (34, 42)

for x in crushed:
    cc = []
    for y in crushed[x]:
        cc.append (y * 1000)
    crushed[x] = cc

branches_unfit = {
    'musanze': retrieve_line(53),
    'rubavu': retrieve_line(56),
    'huye': retrieve_line(59),
    'rusizi': retrieve_line(62),
    'rwamagana': retrieve_line(65)
}

branches_not_unfit = {
    'musanze': retrieve_line(52),
    'rubavu': retrieve_line(55),
    'huye': retrieve_line(58),
    'rusizi': retrieve_line(61),
    'rwamagana': retrieve_line(64)
}

main_cash_desk = retrieve_line(49)

cs_fit = retrieve_line(44)
cs_unfit = retrieve_line(45)
cs_tobe_sorted = retrieve_line(46)
cs_untreated = retrieve_line(47)

# print (json.dumps(printed, indent=4))
# print (json.dumps(uncirculated, indent=4))
# print (json.dumps(crushed, indent=4))
# print (json.dumps(net, indent=4))


wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/HQ Daily Cash Situation report 01.09.2021.xlsx", data_only = True)
sheet_ranges = wb['Cash situation report']
hq = {
    'fit': retrieve_line_hq (9),
    'unfit': retrieve_line_hq (10),
    'mixed': retrieve_line_hq (11),
    'untreated': retrieve_line_hq (12),
    'new': retrieve_line_hq (13),
    'crushed': retrieve_column_hq('I')
}

# print (json.dumps(hq, indent=4))
# print (json.dumps(uncirculated, indent=4))

recentSeries = [ "2003-2004", "2009", "2009", "2009", "2011", "2007", "2019", "2019", "2014", "2014" ]

i = 0
for x in recentSeries:
    uncirculated[x][i] = hq['new'][i]
    i = i + 1
    
i = 0
for x in recentSeries:
    if x in crushed:
        crushed[x][i] = crushed[x][i] + hq['crushed'][i]
    i = i + 1
        
# print (json.dumps(uncirculated, indent=4))

from_branches_not_unfit = {}
from_branches_unfit = {}

wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/Musanze Cash Balance Situation 01.09.2021.xlsx", data_only = True)
sheet_ranges = wb['01-09-2021']
from_branches_not_unfit['musanze'] = retrieve_range_branch (16, 16)
from_branches_unfit['musanze'] = retrieve_range_branch (8, 8)
for x in range(0, len(from_branches_not_unfit['musanze'])):
    from_branches_not_unfit['musanze'][x] = from_branches_not_unfit['musanze'][x] - from_branches_unfit['musanze'][x]

wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/Rwamagana Cash Balance Situation 01.09.2021.xlsx", data_only = True)
sheet_ranges = wb['1-09-2021']
from_branches_not_unfit['rwamagana'] = retrieve_range_branch (15, 15)
from_branches_unfit['rwamagana'] = retrieve_range_branch (8, 8)
for x in range(0, len(from_branches_not_unfit['rwamagana'])):
    from_branches_not_unfit['rwamagana'][x] = from_branches_not_unfit['rwamagana'][x] - from_branches_unfit['rwamagana'][x]

wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/Book1.xlsx", data_only = True)
sheet_ranges = wb['1 09 2021 FRW']
from_branches_not_unfit['huye'] = retrieve_range_branch (14, 14)
from_branches_unfit['huye'] = retrieve_range_branch (8, 8)
for x in range(0, len(from_branches_not_unfit['huye'])):
    from_branches_not_unfit['huye'][x] = from_branches_not_unfit['huye'][x] - from_branches_unfit['huye'][x]

wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/Rubavu Cash situation of 01 September 2021.xlsx", data_only = True)
sheet_ranges = wb['01 September 2021Rwf']
from_branches_not_unfit['rubavu'] = retrieve_range_branch (15, 15, 'B', 'K', '5')
from_branches_unfit['rubavu'] = retrieve_range_branch (8, 9, 'B', 'K', '5')
for x in range(0, len(from_branches_not_unfit['rubavu'])):
    from_branches_not_unfit['rubavu'][x] = from_branches_not_unfit['rubavu'][x] - from_branches_unfit['rubavu'][x]

wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/Rusizi Cash Balance Situation 01.09.2021.xlsx", data_only = True)
sheet_ranges = wb['01-09-2021 (7)']
from_branches_not_unfit['rusizi'] = retrieve_range_branch (17, 17, 'B', 'K', '5')
from_branches_unfit['rusizi'] = retrieve_range_branch (9, 9, 'B', 'K', '5')
for x in range(0, len(from_branches_not_unfit['rusizi'])):
    from_branches_not_unfit['rusizi'][x] = from_branches_not_unfit['rusizi'][x] - from_branches_unfit['rusizi'][x]

# print (json.dumps(from_branches_unfit, indent=4))
# print (json.dumps(from_branches_not_unfit, indent=4))




wb = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/CAISSE RWF Laila 01.09.2021.xlsx", data_only = True)
sheet_ranges = wb['Summary Report']
retrieve_laila = retrieve_laila ()
# print (json.dumps(retrieve_laila, indent=4))

def calculateDemonetized (series, all, current, input):
    for x in series:
        if all[x] > current[x]:
            current[x] = current[x] + input

def calculateUncirculated (series, current, input):
    for x in series:
        if current[x] > 0:
            current[x] = current[x] - input

def copyRows (input, outputRow):
    i = ord ('D')
    for x in input:
        ws[chr(i) + str(outputRow)] = float(x) / 1000
        i = i + 1

def copyRange (input, outputRow1):
    i = outputRow1
    for x in input:
        j = ord ('D')
        for y in input[x]:
            if y != 0:
                ws[chr(j) + str(i)] = float(y) / 1000
            j = j + 1
        i = i + 1

def updateRows ():
    copyRows (hq['fit'], 44)
    copyRows (hq['unfit'], 45)
    copyRows (hq['mixed'], 46)
    copyRows (hq['untreated'], 47)

    copyRows (retrieve_laila, 49)
    
    copyRows (from_branches_not_unfit['musanze'], 52)
    copyRows (from_branches_unfit['musanze'], 53)
    copyRows (from_branches_not_unfit['rubavu'], 55)
    copyRows (from_branches_unfit['rubavu'], 56)
    copyRows (from_branches_not_unfit['huye'], 58)
    copyRows (from_branches_unfit['huye'], 59)
    copyRows (from_branches_not_unfit['rusizi'], 61)
    copyRows (from_branches_unfit['rusizi'], 62)
    copyRows (from_branches_not_unfit['rwamagana'], 64)
    copyRows (from_branches_unfit['rwamagana'], 65)
    
    copyRange (uncirculated, 15)
    copyRange (crushed, 25)

def clearRows ():
    clr = [ 0, 0, 0, 0, 0, 0, 0, 0, 0, 0 ]

    copyRows (clr, 44)
    copyRows (clr, 45)
    copyRows (clr, 46)
    copyRows (clr, 47)

    copyRows (clr, 49)

    copyRows (clr, 52)
    copyRows (clr, 53)
        
wbx = load_workbook(filename = "C:/Users/NTAZINDA/Downloads/C I C 31 08 2021.xlsx")
sheet_ranges = wbx['Curency in circulation']

wb = Workbook()
# ws = wb.active
ws = wb.create_sheet('new')

from copy import copy

for row in sheet_ranges.rows:
    for cell in row:
        new_cell = ws.cell(row=cell.row, column=cell.column, value= cell.value)
        # a = a + 1
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
            

for x in range(1, 77):
    for y in char_range ('A', 'N'):
        ws[y + str(x)] = out[str(x)][y].value

# ws['B73'] = out['72']['B'].value
        
for y in char_range ('A', 'N'):
    ws.column_dimensions[y].width = sheet_ranges.column_dimensions[y].width

ws.freeze_panes = ws['D3']

updateRows ()

wb.save("C I C 01 09 2021.xlsx")